from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os, logging, csv, io, json, uuid, random
from contextlib import asynccontextmanager
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy.orm import DeclarativeBase, mapped_column
from sqlalchemy import String, Float, Integer, Boolean, Text, DateTime, select, func, desc, or_
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import Lead, City, Segment
from genai_helper import call_genai
import pipeline_stages as ps
from pipeline_stages import SERP_API_KEY as _SERP_API_KEY


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

DATABASE_URL = os.getenv("DATABASE_URL")

engine = create_async_engine(DATABASE_URL, echo=False)
AsyncSessionLocal = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ─── ORM MODELS ──────────────────────────────────────────────────────────────

class Base(DeclarativeBase):
    pass


class LeadModel(Base):
    __tablename__ = "leads"

    id = mapped_column(String, primary_key=True)
    business_name = mapped_column(String, nullable=False, default="")
    segment = mapped_column(String, default="Restaurant")
    city = mapped_column(String, nullable=False, default="")
    state = mapped_column(String, default="")
    tier = mapped_column(Integer, default=1)
    address = mapped_column(String, default="")
    phone = mapped_column(String, default="")
    email = mapped_column(String, default="")
    website = mapped_column(String, default="")
    rating = mapped_column(Float, default=0.0)
    num_outlets = mapped_column(Integer, default=1)
    decision_maker_name = mapped_column(String, default="")
    decision_maker_role = mapped_column(String, default="")
    decision_maker_linkedin = mapped_column(String, default="")
    has_dessert_menu = mapped_column(Boolean, default=False)
    hotel_category = mapped_column(String, default="")
    is_chain = mapped_column(Boolean, default=False)
    source = mapped_column(String, default="manual")
    monthly_volume_estimate = mapped_column(String, default="")
    ai_score = mapped_column(Integer, default=0)
    ai_reasoning = mapped_column(Text, default="")
    priority = mapped_column(String, default="Low")
    status = mapped_column(String, default="new")
    created_at = mapped_column(DateTime(timezone=True), nullable=True)
    updated_at = mapped_column(DateTime(timezone=True), nullable=True)


class OutreachEmailModel(Base):
    __tablename__ = "outreach_emails"

    id = mapped_column(String, primary_key=True)
    lead_id = mapped_column(String, nullable=False, default="")
    lead_name = mapped_column(String, default="")
    lead_city = mapped_column(String, default="")
    lead_segment = mapped_column(String, default="")
    subject = mapped_column(String, default="")
    body = mapped_column(Text, default="")
    status = mapped_column(String, default="draft")
    generated_at = mapped_column(DateTime(timezone=True), nullable=True)
    sent_at = mapped_column(DateTime(timezone=True), nullable=True)


def model_to_dict(obj) -> dict:
    result = {}
    for c in obj.__table__.columns:
        val = getattr(obj, c.name)
        if isinstance(val, datetime):
            val = val.isoformat()
        result[c.name] = val
    return result


# ─── LIFESPAN ────────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield
    await engine.dispose()


# ─── APP ─────────────────────────────────────────────────────────────────────

app = FastAPI(lifespan=lifespan)
api_router = APIRouter(prefix="/api")


# ─── SCORING ENGINE ──────────────────────────────────────────────────────────

def calculate_lead_score(data: dict):
    score, reasons = 0, []
    hotel_cat = data.get('hotel_category', '')
    if hotel_cat == '5-star':   score += 30; reasons.append('5-star hotel (+30)')
    elif hotel_cat == '4-star': score += 20; reasons.append('4-star hotel (+20)')
    elif hotel_cat == '3-star': score += 10; reasons.append('3-star hotel (+10)')

    segment = data.get('segment', '')
    seg_pts = {
        # High-volume daily consumers
        'Mithai': 30, 'Bakery': 28, 'FoodProcessing': 26,
        'IceCream': 24, 'Beverage': 22,
        # Medium-volume
        'Catering': 20, 'Cafe': 20, 'Organic': 18,
        'CloudKitchen': 18, 'Brewery': 16,
        # Lower per-unit
        'Restaurant': 15, 'Hotel': 12,
    }
    if segment in seg_pts:
        pts = seg_pts[segment]; score += pts; reasons.append(f'{segment} segment (+{pts})')

    if data.get('is_chain'): score += 15; reasons.append('Chain business (+15)')

    outlets = int(data.get('num_outlets', 1) or 1)
    if outlets >= 10:   score += 15; reasons.append(f'{outlets} outlets (+15)')
    elif outlets >= 3:  score += 10; reasons.append(f'{outlets} outlets (+10)')

    rating = float(data.get('rating', 0) or 0)
    if rating >= 4.5:   score += 10; reasons.append('Rating 4.5+ (+10)')
    elif rating >= 4.0: score += 7;  reasons.append('Rating 4.0+ (+7)')

    tier = int(data.get('tier', 3) or 3)
    if tier == 1:   score += 10; reasons.append('Metro city (+10)')
    elif tier == 2: score += 5;  reasons.append('Tier 2 city (+5)')

    if data.get('has_dessert_menu'): score += 15; reasons.append('Has dessert/sweet menu (+15)')
    if data.get('decision_maker_linkedin'): score += 10; reasons.append('Decision maker on LinkedIn (+10)')

    score = min(score, 100)
    priority = 'High' if score >= 70 else ('Medium' if score >= 40 else 'Low')
    reasoning = ' | '.join(reasons) if reasons else 'Insufficient data for scoring'
    return score, priority, reasoning


def make_lead_obj(data: dict, status: str = "new") -> Lead:
    score, priority, reasoning = calculate_lead_score(data)
    return Lead(
        id=str(uuid.uuid4()),
        business_name=data.get('business_name', ''),
        segment=data.get('segment', 'Restaurant'),
        city=data.get('city', ''),
        state=data.get('state', ''),
        tier=int(data.get('tier', 1) or 1),
        address=data.get('address', ''),
        phone=data.get('phone', ''),
        email=data.get('email', ''),
        website=data.get('website', ''),
        rating=float(data.get('rating', 0) or 0),
        num_outlets=int(data.get('num_outlets', 1) or 1),
        decision_maker_name=data.get('decision_maker_name', ''),
        decision_maker_role=data.get('decision_maker_role', ''),
        decision_maker_linkedin=data.get('decision_maker_linkedin', ''),
        has_dessert_menu=bool(data.get('has_dessert_menu', False)),
        hotel_category=data.get('hotel_category', ''),
        is_chain=bool(data.get('is_chain', False)),
        source=data.get('source', 'manual'),
        monthly_volume_estimate=data.get('monthly_volume_estimate', ''),
        ai_score=score,
        ai_reasoning=reasoning,
        priority=priority,
        status=status,
    )


# ─── PYDANTIC MODELS ─────────────────────────────────────────────────────────

class LeadCreate(BaseModel):
    business_name: str
    segment: str = "Restaurant"
    city: str
    state: str = ""
    tier: int = 1
    address: str = ""
    phone: str = ""
    email: str = ""
    website: str = ""
    rating: float = 0.0
    num_outlets: int = 1
    decision_maker_name: str = ""
    decision_maker_role: str = ""
    decision_maker_linkedin: str = ""
    has_dessert_menu: bool = False
    hotel_category: str = ""
    is_chain: bool = False
    source: str = "manual"
    monthly_volume_estimate: str = ""


class LeadStatusUpdate(BaseModel):
    status: str


class BulkCreateRequest(BaseModel):
    leads: List[dict]


class DiscoverRequest(BaseModel):
    city: str
    segment: str
    state: str = ""


class CityCreate(BaseModel):
    name: str
    state: str = ""
    country: str = "India"
    priority: int = 1


class CityPriorityUpdate(BaseModel):
    priority: int


class SegmentCreate(BaseModel):
    key: str
    label: str = ""
    cluster: str = ""
    description: str = ""
    color: str = "#5C736A"
    priority: int = 1


class SegmentPriorityUpdate(BaseModel):
    priority: int



# ─── API ROUTES ───────────────────────────────────────────────────────────────

@api_router.get("/")
async def root():
    return {"message": "Dhampur Green HORECA Lead Intelligence API v2.0 (PostgreSQL)"}


@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    async with AsyncSessionLocal() as session:
        total_leads = (await session.execute(
            select(func.count()).select_from(LeadModel)
        )).scalar() or 0

        high_priority = (await session.execute(
            select(func.count()).select_from(LeadModel).where(LeadModel.priority == "High")
        )).scalar() or 0

        week_ago = datetime.now(timezone.utc) - timedelta(days=7)
        new_this_week = (await session.execute(
            select(func.count()).select_from(LeadModel).where(LeadModel.created_at >= week_ago)
        )).scalar() or 0

        converted = (await session.execute(
            select(func.count()).select_from(LeadModel).where(LeadModel.status == "converted")
        )).scalar() or 0

        conversion_rate = round((converted / total_leads * 100), 1) if total_leads > 0 else 0

        city_result = await session.execute(
            select(LeadModel.city, func.count(LeadModel.id).label('count'))
            .group_by(LeadModel.city)
            .order_by(desc(func.count(LeadModel.id)))
            .limit(8)
        )
        city_dist = [{"city": row.city or "Unknown", "count": row.count} for row in city_result]

        seg_result = await session.execute(
            select(LeadModel.segment, func.count(LeadModel.id).label('count'))
            .group_by(LeadModel.segment)
            .order_by(desc(func.count(LeadModel.id)))
        )
        seg_dist = [{"segment": row.segment or "Unknown", "count": row.count} for row in seg_result]

        status_result = await session.execute(
            select(LeadModel.status, func.count(LeadModel.id).label('count'))
            .group_by(LeadModel.status)
        )
        status_dist = [{"status": row.status or "Unknown", "count": row.count} for row in status_result]

        recent_result = await session.execute(
            select(LeadModel).order_by(desc(LeadModel.created_at)).limit(6)
        )
        recent = [model_to_dict(r) for r in recent_result.scalars()]

        top_result = await session.execute(
            select(LeadModel).order_by(desc(LeadModel.ai_score)).limit(5)
        )
        top_leads = [model_to_dict(r) for r in top_result.scalars()]

    return {
        "total_leads": total_leads or 0,
        "high_priority": high_priority or 0,
        "new_this_week": new_this_week or 0,
        "converted": converted or 0,
        "conversion_rate": conversion_rate,
        "city_distribution": city_dist,
        "segment_distribution": seg_dist,
        "status_distribution": status_dist,
        "recent_leads": recent,
        "top_leads": top_leads
    }


@api_router.get("/leads")
async def get_leads(
    city: Optional[str] = None, segment: Optional[str] = None,
    priority: Optional[str] = None, status: Optional[str] = None,
    min_score: Optional[int] = None, search: Optional[str] = None,
    limit: int = 100, skip: int = 0,
):
    async with AsyncSessionLocal() as session:
        stmt = select(LeadModel)
        count_stmt = select(func.count()).select_from(LeadModel)

        if city:
            stmt = stmt.where(LeadModel.city.ilike(f"%{city}%"))
            count_stmt = count_stmt.where(LeadModel.city.ilike(f"%{city}%"))
        if segment:
            stmt = stmt.where(LeadModel.segment == segment)
            count_stmt = count_stmt.where(LeadModel.segment == segment)
        if priority:
            stmt = stmt.where(LeadModel.priority == priority)
            count_stmt = count_stmt.where(LeadModel.priority == priority)
        if status:
            stmt = stmt.where(LeadModel.status == status)
            count_stmt = count_stmt.where(LeadModel.status == status)
        if min_score is not None:
            stmt = stmt.where(LeadModel.ai_score >= min_score)
            count_stmt = count_stmt.where(LeadModel.ai_score >= min_score)
        if search:
            search_filter = or_(
                LeadModel.business_name.ilike(f"%{search}%"),
                LeadModel.city.ilike(f"%{search}%"),
                LeadModel.decision_maker_name.ilike(f"%{search}%")
            )
            stmt = stmt.where(search_filter)
            count_stmt = count_stmt.where(search_filter)

        total = (await session.execute(count_stmt)).scalar() or 0
        result = await session.execute(
            stmt.order_by(desc(LeadModel.ai_score)).offset(skip).limit(limit)
        )
        leads = [model_to_dict(r) for r in result.scalars()]

    return {"leads": leads, "total": total}


@api_router.get("/leads/csv-template")
async def get_csv_template():
    """Generate an Excel template (.xlsx) with in-cell dropdown validation
    for the city and segment columns, sourced from the active pipeline config."""
    import io as _io

    # ── Fetch active cities & segments from DB ───────────────────────────────
    async with AsyncSessionLocal() as session:
        city_rows = (await session.execute(
            select(City).where(City.is_active == True).order_by(City.priority.asc(), City.name.asc())
        )).scalars().all()
        seg_rows = (await session.execute(
            select(Segment).where(Segment.is_active == True).order_by(Segment.priority.asc(), Segment.label.asc())
        )).scalars().all()

    city_names = [c.name for c in city_rows] or ["Mumbai", "Delhi", "Bangalore"]
    seg_keys   = [s.key  for s in seg_rows]  or ["Hotel", "Restaurant", "Cafe", "Bakery"]

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Hidden reference sheet for dropdown lists ────────────────────────────
    ref_ws = wb.create_sheet("_ref", 0)
    ref_ws.sheet_state = "hidden"
    for i, name in enumerate(city_names, start=1):
        ref_ws.cell(row=i, column=1, value=name)
    for i, key in enumerate(seg_keys, start=1):
        ref_ws.cell(row=i, column=2, value=key)

    city_ref = f"_ref!$A$1:$A${len(city_names)}"
    seg_ref  = f"_ref!$B$1:$B${len(seg_keys)}"

    # ── Main data sheet ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Leads", 1)
    wb.active = ws

    COLUMNS = [
        ("business_name",          "Business Name *"),
        ("segment",                "Segment *"),
        ("city",                   "City *"),
        ("state",                  "State"),
        ("tier",                   "Tier (1/2/3)"),
        ("address",                "Address"),
        ("phone",                  "Phone"),
        ("email",                  "Email"),
        ("website",                "Website"),
        ("rating",                 "Rating (0-5)"),
        ("num_outlets",            "No. of Outlets"),
        ("decision_maker_name",    "Decision Maker Name"),
        ("decision_maker_role",    "Decision Maker Role"),
        ("decision_maker_linkedin","Decision Maker LinkedIn"),
        ("has_dessert_menu",       "Has Dessert Menu (true/false)"),
        ("hotel_category",         "Hotel Category"),
        ("is_chain",               "Is Chain (true/false)"),
        ("monthly_volume_estimate","Monthly Volume Estimate"),
    ]

    SAMPLE = [
        "The Grand Palace Hotel", seg_keys[0], city_names[0],
        city_rows[0].state if city_rows else "",
        "1", "Colaba, Mumbai", "+91-9876543210",
        "procurement@grandhotel.com", "www.grandhotel.com",
        "4.5", "3", "Rajesh Kumar", "Procurement Manager",
        "linkedin.com/in/rajeshkumar", "true", "5-star", "false", "500-800 kg",
    ]

    # Header style
    hdr_fill   = PatternFill("solid", fgColor="627F31")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="CCCCCC")
    thin_border= Border(left=thin_side, right=thin_side, bottom=thin_side)

    for col_idx, (key, label) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        # Header row
        hdr_cell = ws.cell(row=1, column=col_idx, value=label)
        hdr_cell.fill   = hdr_fill
        hdr_cell.font   = hdr_font
        hdr_cell.alignment = hdr_align
        hdr_cell.border = thin_border
        # Sample row
        sample_cell = ws.cell(row=2, column=col_idx, value=SAMPLE[col_idx - 1])
        sample_cell.alignment = Alignment(vertical="center")
        # Column widths — wider for free-text fields
        wide_cols = {"address": 40, "website": 32, "email": 32,
                     "decision_maker_linkedin": 36, "monthly_volume_estimate": 24}
        ws.column_dimensions[col_letter].width = wide_cols.get(key, max(len(label), 18))

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"  # freeze header

    # ── Dropdown validations (rows 2-1001 = 1000 data rows) ──────────────────
    MAX_ROW = 1001

    seg_col  = next(i for i, (k, _) in enumerate(COLUMNS, 1) if k == "segment")
    city_col = next(i for i, (k, _) in enumerate(COLUMNS, 1) if k == "city")

    dv_seg = DataValidation(
        type="list",
        formula1=seg_ref,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid segment",
        error=f"Choose one of the active segments: {', '.join(seg_keys[:8])}{'…' if len(seg_keys) > 8 else ''}",
    )
    dv_seg.sqref = f"{get_column_letter(seg_col)}2:{get_column_letter(seg_col)}{MAX_ROW}"
    ws.add_data_validation(dv_seg)

    dv_city = DataValidation(
        type="list",
        formula1=city_ref,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid city",
        error=f"Choose one of the active target cities: {', '.join(city_names[:8])}{'…' if len(city_names) > 8 else ''}",
    )
    dv_city.sqref = f"{get_column_letter(city_col)}2:{get_column_letter(city_col)}{MAX_ROW}"
    ws.add_data_validation(dv_city)

    # ── Boolean dropdowns (true / false) ─────────────────────────────────────
    bool_cols = [k for k, _ in COLUMNS if k in ("has_dessert_menu", "is_chain")]
    dv_bool = DataValidation(
        type="list",
        formula1='"true,false"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="Choose true or false",
    )
    for bool_key in bool_cols:
        col_idx = next(i for i, (k, _) in enumerate(COLUMNS, 1) if k == bool_key)
        col_letter = get_column_letter(col_idx)
        dv_bool.sqref = f"{col_letter}2:{col_letter}{MAX_ROW}" if not dv_bool.sqref else f"{dv_bool.sqref} {col_letter}2:{col_letter}{MAX_ROW}"
    ws.add_data_validation(dv_bool)

    # ── Tier dropdown (1 / 2 / 3) ────────────────────────────────────────────
    tier_col_idx = next(i for i, (k, _) in enumerate(COLUMNS, 1) if k == "tier")
    tier_letter = get_column_letter(tier_col_idx)
    dv_tier = DataValidation(
        type="list",
        formula1='"1,2,3"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid tier",
        error="Choose 1 (Metro), 2 (Tier 2), or 3 (Tier 3)",
    )
    dv_tier.sqref = f"{tier_letter}2:{tier_letter}{MAX_ROW}"
    ws.add_data_validation(dv_tier)

    # ── Hotel category dropdown ───────────────────────────────────────────────
    hcat_col_idx = next((i for i, (k, _) in enumerate(COLUMNS, 1) if k == "hotel_category"), None)
    if hcat_col_idx:
        hcat_letter = get_column_letter(hcat_col_idx)
        dv_hcat = DataValidation(
            type="list",
            formula1='"3-star,4-star,5-star"',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid category",
            error="Choose 3-star, 4-star, or 5-star (leave blank for non-hotels)",
        )
        dv_hcat.sqref = f"{hcat_letter}2:{hcat_letter}{MAX_ROW}"
        ws.add_data_validation(dv_hcat)

    # ── Reference sheet visible to user ──────────────────────────────────────
    info_ws = wb.create_sheet("Valid Values", 2)
    info_ws.column_dimensions["A"].width = 24
    info_ws.column_dimensions["B"].width = 24

    def write_section(sheet, start_row, title, values, col):
        hdr = sheet.cell(row=start_row, column=col, value=title)
        hdr.font = Font(bold=True, color="FFFFFF", size=10)
        hdr.fill = PatternFill("solid", fgColor="627F31")
        hdr.alignment = Alignment(horizontal="center")
        for i, v in enumerate(values, start=start_row + 1):
            sheet.cell(row=i, column=col, value=v)

    write_section(info_ws, 1, f"Active Cities ({len(city_names)})", city_names, 1)
    write_section(info_ws, 1, f"Active Segments ({len(seg_keys)})",  seg_keys,   2)

    # ── Serialize & return ───────────────────────────────────────────────────
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=horeca_leads_template.xlsx"},
    )


@api_router.post("/leads/upload-csv")
async def upload_csv(file: UploadFile = File(...)):
    content = await file.read()
    try:
        text = content.decode('utf-8-sig')
    except Exception:
        text = content.decode('latin-1')

    reader = csv.DictReader(io.StringIO(text))
    created = []
    errors = []

    async with AsyncSessionLocal() as session:
        for i, row in enumerate(reader):
            try:
                lead_data = {
                    "business_name": str(row.get('business_name', '')).strip(),
                    "segment": str(row.get('segment', 'Restaurant')).strip() or 'Restaurant',
                    "city": str(row.get('city', '')).strip(),
                    "state": str(row.get('state', '')).strip(),
                    "tier": int(str(row.get('tier', '1')).strip() or '1'),
                    "address": str(row.get('address', '')).strip(),
                    "phone": str(row.get('phone', '')).strip(),
                    "email": str(row.get('email', '')).strip(),
                    "website": str(row.get('website', '')).strip(),
                    "rating": float(str(row.get('rating', '0')).strip() or '0'),
                    "num_outlets": int(str(row.get('num_outlets', '1')).strip() or '1'),
                    "decision_maker_name": str(row.get('decision_maker_name', '')).strip(),
                    "decision_maker_role": str(row.get('decision_maker_role', '')).strip(),
                    "decision_maker_linkedin": str(row.get('decision_maker_linkedin', '')).strip(),
                    "has_dessert_menu": str(row.get('has_dessert_menu', 'false')).lower().strip() in ('true', '1', 'yes'),
                    "hotel_category": str(row.get('hotel_category', '')).strip(),
                    "is_chain": str(row.get('is_chain', 'false')).lower().strip() in ('true', '1', 'yes'),
                    "source": "csv_upload",
                    "monthly_volume_estimate": str(row.get('monthly_volume_estimate', '')).strip()
                }
                if not lead_data['business_name'] or not lead_data['city']:
                    errors.append(f"Row {i + 2}: Missing business_name or city")
                    continue
                doc = make_lead_obj(lead_data)
                session.add(doc)
                created.append(doc.to_dict())
            except Exception as e:
                errors.append(f"Row {i + 2}: {str(e)}")
        await session.commit()

    return {"created": len(created), "errors": errors}


@api_router.post("/leads/discover")
async def discover_leads(req: DiscoverRequest):
    city    = req.city.strip()
    segment = req.segment.strip()
    state   = req.state.strip()

    logger.info(f"[DISCOVER] Request — city={city!r}, segment={segment!r}, state={state!r}")

    # ───────────────────────────────────────────────────────────────────
    # PATH A — Real pipeline (SerpAPI available)
    # Runs stages 1–4: Extract → AI enrich → KPI filter → Dedup
    # Returns leads for user review; saving happens via /leads/bulk-create
    # ───────────────────────────────────────────────────────────────────
    if _SERP_API_KEY:
        try:
            async with AsyncSessionLocal() as session:
                # Stage 1: SerpAPI Google Maps — filtered to the chosen segment
                raw = await ps.extract_business_data(city, session)
                logger.info(f"[DISCOVER] Stage 1 extracted {len(raw)} raw places")

                if not raw:
                    logger.warning(f"[DISCOVER] No SerpAPI results for city={city!r} segment={segment!r}")
                    return []

                # Stage 2: Gemini AI enrichment
                enriched = await ps.ai_process_business_data(raw, session)
                logger.info(f"[DISCOVER] Stage 2 AI-enriched {len(enriched)} businesses")

                # Stage 3: KPI scoring + filtering
                filtered = await ps.apply_kpi_filtering(enriched, session)
                logger.info(f"[DISCOVER] Stage 3 KPI filtered → {len(filtered)} passed")

                # Stage 4: Dedup against DB (skip businesses already stored)
                deduped = await ps.deduplicate_leads(filtered, session)
                logger.info(f"[DISCOVER] Stage 4 dedup → {len(deduped)} unique leads")

            # Normalise output to the shape the frontend expects
            results = []
            for biz in deduped:
                results.append({
                    "business_name":           biz.get("business_name", ""),
                    "segment":                 biz.get("segment", segment),
                    "city":                    biz.get("city", city),
                    "state":                   biz.get("state", state),
                    "tier":                    biz.get("tier", 1),
                    "address":                 biz.get("address", ""),
                    "phone":                   biz.get("phone", ""),
                    "email":                   biz.get("email", ""),
                    "website":                 biz.get("website", ""),
                    "rating":                  biz.get("rating", 0.0),
                    "num_outlets":             biz.get("num_outlets", 1),
                    "hotel_category":          biz.get("hotel_category", ""),
                    "is_chain":                bool(biz.get("is_chain", False)),
                    "has_dessert_menu":        bool(biz.get("has_dessert_menu", False)),
                    "decision_maker_name":     biz.get("decision_maker_name", ""),
                    "decision_maker_role":     biz.get("decision_maker_role", ""),
                    "decision_maker_linkedin": biz.get("decision_maker_linkedin", ""),
                    "monthly_volume_estimate": f"{biz.get('monthly_sugar_estimate_kg', '')} kg".strip(),
                    "ai_score":                int(biz.get("kpi_score", 0) or 0),
                    "ai_reasoning":            biz.get("ai_reasoning", ""),
                    "priority":                biz.get("priority", "Low"),
                    "source":                  biz.get("source", "serpapi_google_maps"),
                })
            logger.info(f"[DISCOVER] Returning {len(results)} leads to frontend")
            return results

        except Exception as exc:
            logger.error(f"[DISCOVER] Pipeline failed: {exc}", exc_info=True)
            return []

    logger.warning("[DISCOVER] SERP_API_KEY not configured — returning empty results")
    return []


@api_router.post("/leads/bulk-create")
async def bulk_create_leads(req: BulkCreateRequest):
    created = []
    async with AsyncSessionLocal() as session:
        for lead_dict in req.leads:
            lead_dict.pop('ai_score', None)
            lead_dict.pop('ai_reasoning', None)
            lead_dict.pop('priority', None)
            doc = make_lead_obj(lead_dict)
            session.add(doc)
            created.append(doc.to_dict())
        await session.commit()
    return {"created": len(created), "leads": created}


@api_router.post("/leads")
async def create_lead(lead: LeadCreate):
    doc = make_lead_obj(lead.model_dump())
    async with AsyncSessionLocal() as session:
        session.add(doc)
        await session.commit()
    return doc.to_dict()


@api_router.get("/leads/{lead_id}")
async def get_lead(lead_id: str):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(LeadModel).where(LeadModel.id == lead_id))
        lead = result.scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        return model_to_dict(lead)


@api_router.put("/leads/{lead_id}/status")
async def update_lead_status(lead_id: str, body: LeadStatusUpdate):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(LeadModel).where(LeadModel.id == lead_id))
        lead = result.scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        lead.status = body.status
        lead.updated_at = datetime.now(timezone.utc)
        await session.commit()
        await session.refresh(lead)
        return model_to_dict(lead)


@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(LeadModel).where(LeadModel.id == lead_id))
        lead = result.scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        await session.delete(lead)
        await session.commit()
    return {"message": "Lead deleted"}


@api_router.post("/leads/{lead_id}/qualify-ai")
async def qualify_lead_ai(lead_id: str):
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(LeadModel).where(LeadModel.id == lead_id)
        )
        lead_obj = result.scalar_one_or_none()

        if not lead_obj:
            raise HTTPException(status_code=404, detail="Lead not found")

        lead = model_to_dict(lead_obj)

    try:
        prompt = f"""Qualify this HORECA business for Dhampur Green (sugar/jaggery supplier):
Business: {lead['business_name']}, Segment: {lead['segment']}
Location: {lead['city']}, {lead.get('state') or 'India'} | Rating: {lead['rating']}/5 | Outlets: {lead['num_outlets']}
Hotel Category: {lead.get('hotel_category') or 'N/A'} | Dessert Menu: {lead['has_dessert_menu']} | Chain: {lead['is_chain']}

Respond ONLY with valid JSON:
{{"ai_score":<0-100>,"monthly_volume_kg":"<range>","qualification_summary":"<2-3 sentences>","sugar_use_cases":["<uc1>","<uc2>","<uc3>"],"key_insight":"<sales insight>","priority":"<High/Medium/Low>","best_contact_time":"<recommendation>"}}"""

        response = call_genai(prompt, force_json=True)

        json_str = response.strip()

        # Clean markdown if present
        if '```json' in json_str:
            json_str = json_str.split('```json')[1].split('```')[0]
        elif '```' in json_str:
            json_str = json_str.split('```')[1].split('```')[0]

        try:
            ai_data = json.loads(json_str.strip())
        except json.JSONDecodeError:
            raise HTTPException(status_code=500, detail="Invalid AI JSON response")

        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(LeadModel).where(LeadModel.id == lead_id)
            )
            lead_obj = result.scalar_one_or_none()

            if lead_obj:
                lead_obj.ai_score = int(ai_data.get('ai_score', lead.get('ai_score', 0)))
                lead_obj.ai_reasoning = ai_data.get('qualification_summary', '')
                lead_obj.priority = ai_data.get('priority', 'Medium')
                lead_obj.monthly_volume_estimate = ai_data.get('monthly_volume_kg', '')
                lead_obj.updated_at = datetime.now(timezone.utc)

                await session.commit()
                await session.refresh(lead_obj)

                updated = model_to_dict(lead_obj)

        return {
            "lead": updated,
            "ai_analysis": ai_data
        }

    except Exception as e:
        logger.error(f"AI qualify error: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"AI qualification failed: {str(e)}"
        )


@api_router.post("/leads/{lead_id}/generate-email")
async def generate_email(lead_id: str):
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(LeadModel).where(LeadModel.id == lead_id)
        )
        lead_obj = result.scalar_one_or_none()

        if not lead_obj:
            raise HTTPException(status_code=404, detail="Lead not found")

        lead = model_to_dict(lead_obj)

    try:
        dm = lead.get('decision_maker_name') or 'Procurement Manager'
        first_name = dm.split()[0] if dm else 'Sir/Madam'

        prompt = f"""Write a personalized B2B outreach email for Dhampur Green targeting:
Business: {lead['business_name']} ({lead['segment']}, {lead['city']})
Decision Maker: {dm} ({lead.get('decision_maker_role') or 'F&B Head'})
Rating: {lead['rating']}/5 | Outlets: {lead['num_outlets']} | Dessert Menu: {lead['has_dessert_menu']}
Monthly Volume Estimate: {lead.get('monthly_volume_estimate') or 'Unknown'}

Dhampur Green Products: Premium refined sugar (M30/S30), sulphur-free jaggery, brown sugar, organic cane sugar, khandsari, icing sugar.

Write a 5-7 line professional email with specific subject, personalized opener, value prop, product rec, soft CTA, sign-off from "Arjun Mehta | Regional Sales Manager, Dhampur Green | +91-98765-43210"

Format EXACTLY:
SUBJECT: [subject]

Dear {first_name},
[body]"""

        response = call_genai(prompt)

        lines = response.strip().split('\n')

        subject = ""
        body_lines = []
        past_subject = False

        for line in lines:
            if line.startswith('SUBJECT:') and not past_subject:
                subject = line.replace('SUBJECT:', '').strip()
                past_subject = True
            else:
                body_lines.append(line)

        body = '\n'.join(body_lines).strip()

        now = datetime.now(timezone.utc)

        doc = {
            "id": str(uuid.uuid4()),
            "lead_id": lead_id,
            "lead_name": lead['business_name'],
            "lead_city": lead['city'],
            "lead_segment": lead['segment'],
            "subject": subject,
            "body": body,
            "status": "draft",
            "generated_at": now,
            "sent_at": None
        }

        async with AsyncSessionLocal() as session:
            session.add(OutreachEmailModel(**doc))
            await session.commit()

        doc.pop('sent_at', None)

        return doc

    except Exception as e:
        logger.error(f"Email gen error: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Email generation failed: {str(e)}"
        )
    

@api_router.get("/outreach/emails")
async def get_all_emails(limit: int = 50):
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(OutreachEmailModel).order_by(desc(OutreachEmailModel.generated_at)).limit(limit)
        )
        return [model_to_dict(e) for e in result.scalars()]


@api_router.get("/outreach/{lead_id}/emails")
async def get_lead_emails(lead_id: str):
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(OutreachEmailModel)
            .where(OutreachEmailModel.lead_id == lead_id)
            .order_by(desc(OutreachEmailModel.generated_at))
            .limit(20)
        )
        return [model_to_dict(e) for e in result.scalars()]


@api_router.put("/outreach/{email_id}/mark-sent")
async def mark_email_sent(email_id: str):
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(OutreachEmailModel).where(OutreachEmailModel.id == email_id)
        )
        email = result.scalar_one_or_none()
        if not email:
            raise HTTPException(status_code=404, detail="Email not found")
        email.status = "sent"
        email.sent_at = datetime.now(timezone.utc)
        await session.commit()
        await session.refresh(email)
        return model_to_dict(email)


# ─── CITY MANAGEMENT ────────────────────────────────────────────────────────

@api_router.get("/cities")
async def list_cities():
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(City).order_by(City.priority.asc(), City.name.asc()))
        cities = result.scalars().all()
        return [c.to_dict() for c in cities]


@api_router.post("/cities", status_code=201)
async def add_city(body: CityCreate):
    async with AsyncSessionLocal() as session:
        # Prevent duplicates (case-insensitive)
        existing = await session.execute(
            select(City).where(func.lower(City.name) == body.name.strip().lower())
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail=f"City '{body.name}' already exists")
        city = City(
            name=body.name.strip(),
            state=body.state.strip(),
            country=body.country,
            priority=body.priority,
            is_active=True,
        )
        session.add(city)
        await session.commit()
        await session.refresh(city)
        return city.to_dict()


@api_router.put("/cities/{city_id}/toggle")
async def toggle_city(city_id: int):
    async with AsyncSessionLocal() as session:
        city = await session.get(City, city_id)
        if not city:
            raise HTTPException(status_code=404, detail="City not found")
        city.is_active = not city.is_active
        await session.commit()
        await session.refresh(city)
        return city.to_dict()


@api_router.put("/cities/{city_id}/priority")
async def update_city_priority(city_id: int, body: CityPriorityUpdate):
    async with AsyncSessionLocal() as session:
        city = await session.get(City, city_id)
        if not city:
            raise HTTPException(status_code=404, detail="City not found")
        city.priority = max(1, body.priority)
        await session.commit()
        await session.refresh(city)
        return city.to_dict()


@api_router.delete("/cities/{city_id}", status_code=204)
async def delete_city(city_id: int):
    async with AsyncSessionLocal() as session:
        city = await session.get(City, city_id)
        if not city:
            raise HTTPException(status_code=404, detail="City not found")
        await session.delete(city)
        await session.commit()
    return Response(status_code=204)


# ─── SEGMENT MANAGEMENT ─────────────────────────────────────────────────────

# Master catalog — seeded once, then toggled / re-prioritised by admin
SEGMENT_CATALOG = [
    {"key": "Mithai",        "label": "Mithai / Sweets",    "cluster": "Traditional Sweets",     "color": "#A0522D", "description": "Mithai shops & sweet chains; highest sugar density per kg of product"},
    {"key": "Bakery",        "label": "Bakery",             "cluster": "Bakery & Confectionery",  "color": "#B85C38", "description": "Bakeries, patisseries & cake shops; 15–30% sugar per product batch"},
    {"key": "FoodProcessing","label": "Food Processing",    "cluster": "Food Processing",         "color": "#7B6D47", "description": "Industrial food processors, packaged-food manufacturers"},
    {"key": "IceCream",      "label": "Ice Cream",          "cluster": "Dairy & Frozen",           "color": "#C4878A", "description": "Ice-cream parlours & dairy-frozen chains; 12–18% sugar in mix"},
    {"key": "Beverage",      "label": "Beverage",           "cluster": "Beverage",                "color": "#4A7FA5", "description": "Juice bars, RTD beverage makers & soft-drink producers"},
    {"key": "Catering",      "label": "Catering",           "cluster": "HORECA",                  "color": "#6B5E44", "description": "Event & bulk caterers; large per-event sugar volumes"},
    {"key": "Cafe",          "label": "Café",               "cluster": "HORECA",                  "color": "#8FA39A", "description": "Coffee shops & cafés; syrups, frappes, baked goods"},
    {"key": "CloudKitchen",  "label": "Cloud Kitchen",      "cluster": "HORECA",                  "color": "#D4956A", "description": "Delivery-only kitchens; high-throughput dessert menus"},
    {"key": "Organic",       "label": "Organic",            "cluster": "Health & Organic",         "color": "#5A8A3C", "description": "Organic food brands, health-food stores, natural sweetener buyers"},
    {"key": "Brewery",       "label": "Brewery",            "cluster": "Fermentation",             "color": "#7B4F72", "description": "Craft breweries & fermentation units; sucrose for fermentation"},
    {"key": "Restaurant",    "label": "Restaurant",         "cluster": "HORECA",                  "color": "#3D6B56", "description": "Full-service restaurants & dhabas; dessert & cooking sugar"},
    {"key": "Hotel",         "label": "Hotel",              "cluster": "HORECA",                  "color": "#662B01", "description": "Hotel F&B departments; multiple restaurant, pastry, banquet consumption"},
]


@api_router.get("/segments")
async def list_segments():
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(Segment).order_by(Segment.priority.asc(), Segment.label.asc())
        )
        segs = result.scalars().all()
        return [s.to_dict() for s in segs]


@api_router.post("/segments/seed", status_code=201)
async def seed_segments():
    """Idempotent seed — inserts catalog entries that do not yet exist."""
    async with AsyncSessionLocal() as session:
        created = []
        for i, entry in enumerate(SEGMENT_CATALOG, start=1):
            existing = await session.execute(
                select(Segment).where(func.lower(Segment.key) == entry["key"].lower())
            )
            if existing.scalar_one_or_none():
                continue
            seg = Segment(
                key=entry["key"],
                label=entry["label"],
                cluster=entry["cluster"],
                description=entry["description"],
                color=entry["color"],
                is_active=True,
                priority=i,
            )
            session.add(seg)
            created.append(entry["key"])
        await session.commit()
        return {"seeded": created, "total": len(created)}


@api_router.post("/segments", status_code=201)
async def create_segment(body: SegmentCreate):
    """Create a custom (admin-defined) segment."""
    async with AsyncSessionLocal() as session:
        existing = await session.execute(
            select(Segment).where(func.lower(Segment.key) == body.key.strip().lower())
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail=f"Segment key '{body.key}' already exists")
        # auto-assign priority = current max + 1
        max_priority = (await session.execute(select(func.max(Segment.priority)))).scalar() or 0
        seg = Segment(
            key=body.key.strip(),
            label=body.label.strip() or body.key.strip(),
            cluster=body.cluster.strip(),
            description=body.description.strip(),
            color=body.color,
            is_active=True,
            priority=max_priority + 1,
        )
        session.add(seg)
        await session.commit()
        await session.refresh(seg)
        return seg.to_dict()


@api_router.delete("/segments/{seg_id}", status_code=204)
async def delete_segment(seg_id: int):
    async with AsyncSessionLocal() as session:
        seg = await session.get(Segment, seg_id)
        if not seg:
            raise HTTPException(status_code=404, detail="Segment not found")
        await session.delete(seg)
        await session.commit()
    return Response(status_code=204)


@api_router.put("/segments/{seg_id}/toggle")
async def toggle_segment(seg_id: int):
    async with AsyncSessionLocal() as session:
        seg = await session.get(Segment, seg_id)
        if not seg:
            raise HTTPException(status_code=404, detail="Segment not found")
        seg.is_active = not seg.is_active
        await session.commit()
        await session.refresh(seg)
        return seg.to_dict()


@api_router.put("/segments/{seg_id}/priority")
async def update_segment_priority(seg_id: int, body: SegmentPriorityUpdate):
    async with AsyncSessionLocal() as session:
        seg = await session.get(Segment, seg_id)
        if not seg:
            raise HTTPException(status_code=404, detail="Segment not found")
        seg.priority = max(1, body.priority)
        await session.commit()
        await session.refresh(seg)
        return seg.to_dict()


@api_router.post("/seed-mock-data")
async def seed_mock_data():
    async with AsyncSessionLocal() as session:
        count = (await session.execute(
            select(func.count()).select_from(LeadModel)
        )).scalar() or 0
    if count > 0:
        return {"message": f"Already has {count} leads", "count": count}

    mock_leads = [
        {"business_name": "The Taj Mahal Palace", "segment": "Hotel", "city": "Mumbai", "state": "Maharashtra", "tier": 1, "address": "Apollo Bunder, Colaba, Mumbai", "phone": "+91-22-6665-3366", "email": "fbprocurement@tajhotels.com", "website": "www.tajhotels.com", "rating": 4.8, "num_outlets": 12, "decision_maker_name": "Rakesh Nair", "decision_maker_role": "F&B Procurement Director", "decision_maker_linkedin": "linkedin.com/in/rakesh-nair-taj", "has_dessert_menu": True, "hotel_category": "5-star", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "800-1200 kg"},
        {"business_name": "Grand Hyatt Mumbai", "segment": "Hotel", "city": "Mumbai", "state": "Maharashtra", "tier": 1, "address": "Santacruz East, Mumbai", "phone": "+91-22-6676-1234", "email": "procurement@grandhyattmumbai.com", "website": "www.grandhyatt.com", "rating": 4.6, "num_outlets": 2, "decision_maker_name": "Priya Sharma", "decision_maker_role": "Purchase Manager", "decision_maker_linkedin": "linkedin.com/in/priya-sharma-hyatt", "has_dessert_menu": True, "hotel_category": "5-star", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "600-800 kg"},
        {"business_name": "Monginis Cake Shop Chain", "segment": "Bakery", "city": "Mumbai", "state": "Maharashtra", "tier": 1, "address": "Multiple locations, Mumbai", "phone": "+91-22-2376-5678", "email": "procurement@monginis.net", "website": "www.monginis.net", "rating": 4.2, "num_outlets": 230, "decision_maker_name": "Suhail Khorakiwala", "decision_maker_role": "Procurement Head", "decision_maker_linkedin": "linkedin.com/in/khorakiwala-monginis", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "5000-8000 kg"},
        {"business_name": "La Folie Patisserie", "segment": "Bakery", "city": "Mumbai", "state": "Maharashtra", "tier": 1, "address": "Khar West, Mumbai", "phone": "+91-22-6503-4567", "email": "hello@lafolie.in", "website": "www.lafolie.in", "rating": 4.7, "num_outlets": 6, "decision_maker_name": "Sanjana Patel", "decision_maker_role": "Owner & Head Pastry Chef", "decision_maker_linkedin": "linkedin.com/in/sanjana-lafolie", "has_dessert_menu": True, "hotel_category": "", "is_chain": False, "source": "mock_data", "monthly_volume_estimate": "300-500 kg"},
        {"business_name": "Natural Ice Cream", "segment": "IceCream", "city": "Mumbai", "state": "Maharashtra", "tier": 1, "address": "Juhu, Mumbai", "phone": "+91-22-2618-3456", "email": "supply@naturals.in", "website": "www.naturals.in", "rating": 4.5, "num_outlets": 135, "decision_maker_name": "Raghunandan Kamath", "decision_maker_role": "Owner", "decision_maker_linkedin": "linkedin.com/in/naturals-kamath", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "3000-5000 kg"},
        {"business_name": "Kailash Parbat Mithai", "segment": "Mithai", "city": "Mumbai", "state": "Maharashtra", "tier": 1, "address": "Colaba, Mumbai", "phone": "+91-22-2202-9456", "email": "orders@kailashparbat.com", "website": "www.kailashparbat.in", "rating": 4.3, "num_outlets": 22, "decision_maker_name": "Vijay Gidwani", "decision_maker_role": "Purchase Manager", "decision_maker_linkedin": "linkedin.com/in/kailash-parbat", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "1500-2500 kg"},
        {"business_name": "Smoke House Deli Mumbai", "segment": "Restaurant", "city": "Mumbai", "state": "Maharashtra", "tier": 1, "address": "Bandra West, Mumbai", "phone": "+91-22-6520-3456", "email": "procurement@smokehousedeli.com", "website": "www.smokehousedeli.com", "rating": 4.3, "num_outlets": 8, "decision_maker_name": "Manish Mehrotra", "decision_maker_role": "F&B Director", "decision_maker_linkedin": "linkedin.com/in/smokehouse-procurement", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "200-400 kg"},
        {"business_name": "BOX8 Cloud Kitchen", "segment": "CloudKitchen", "city": "Mumbai", "state": "Maharashtra", "tier": 1, "address": "Andheri East, Mumbai", "phone": "+91-22-7123-8901", "email": "supply@box8.in", "website": "www.box8.in", "rating": 4.0, "num_outlets": 45, "decision_maker_name": "Anshul Gupta", "decision_maker_role": "Supply Chain Manager", "decision_maker_linkedin": "linkedin.com/in/box8-supply-chain", "has_dessert_menu": False, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "500-800 kg"},
        {"business_name": "The Leela Palace Delhi", "segment": "Hotel", "city": "Delhi", "state": "Delhi", "tier": 1, "address": "Diplomatic Enclave, Chanakyapuri", "phone": "+91-11-3933-1234", "email": "procurement@theleela.com", "website": "www.theleela.com", "rating": 4.8, "num_outlets": 8, "decision_maker_name": "Vikram Nair", "decision_maker_role": "F&B Procurement Head", "decision_maker_linkedin": "linkedin.com/in/vikram-nair-leela", "has_dessert_menu": True, "hotel_category": "5-star", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "700-1000 kg"},
        {"business_name": "The Baker's Dozen Delhi", "segment": "Bakery", "city": "Delhi", "state": "Delhi", "tier": 1, "address": "Khan Market, Delhi", "phone": "+91-11-4150-7890", "email": "procurement@bakersdozen.in", "website": "www.thebakersdozen.in", "rating": 4.5, "num_outlets": 28, "decision_maker_name": "Aditi Handa", "decision_maker_role": "CEO & Founder", "decision_maker_linkedin": "linkedin.com/in/aditi-handa-bakers", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "800-1200 kg"},
        {"business_name": "Haldiram's Delhi", "segment": "Mithai", "city": "Delhi", "state": "Delhi", "tier": 1, "address": "Lajpat Nagar, Delhi", "phone": "+91-11-2921-5678", "email": "supply@haldirams.com", "website": "www.haldirams.com", "rating": 4.4, "num_outlets": 150, "decision_maker_name": "Procurement Director", "decision_maker_role": "Procurement Director", "decision_maker_linkedin": "linkedin.com/in/haldirams-procurement", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "10000-20000 kg"},
        {"business_name": "Blue Tokai Coffee Roasters", "segment": "Cafe", "city": "Delhi", "state": "Delhi", "tier": 1, "address": "Saket, Delhi", "phone": "+91-11-4200-5678", "email": "wholesale@bluetokaicoffee.com", "website": "www.bluetokaicoffee.com", "rating": 4.5, "num_outlets": 30, "decision_maker_name": "Matt Chitharanjan", "decision_maker_role": "Founder", "decision_maker_linkedin": "linkedin.com/in/matt-bluetokai", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "200-400 kg"},
        {"business_name": "Punjab Grill Delhi", "segment": "Restaurant", "city": "Delhi", "state": "Delhi", "tier": 1, "address": "Select Citywalk, Saket, Delhi", "phone": "+91-11-4100-7890", "email": "procurement@punjabgrill.in", "website": "www.punjabgrill.in", "rating": 4.2, "num_outlets": 15, "decision_maker_name": "Sanjeev Nanda", "decision_maker_role": "Director Operations", "decision_maker_linkedin": "linkedin.com/in/punjab-grill-ops", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "400-700 kg"},
        {"business_name": "ITC Windsor Bengaluru", "segment": "Hotel", "city": "Bangalore", "state": "Karnataka", "tier": 1, "address": "Windsor Square, Golf Course Road, Bangalore", "phone": "+91-80-2226-9898", "email": "windsor.procurement@itchotels.in", "website": "www.itchotels.in", "rating": 4.7, "num_outlets": 4, "decision_maker_name": "Sanjay Menon", "decision_maker_role": "F&B Manager", "decision_maker_linkedin": "linkedin.com/in/sanjay-itcwindsor", "has_dessert_menu": True, "hotel_category": "5-star", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "500-800 kg"},
        {"business_name": "Third Wave Coffee", "segment": "Cafe", "city": "Bangalore", "state": "Karnataka", "tier": 1, "address": "Indiranagar, Bangalore", "phone": "+91-80-4123-5678", "email": "procurement@thirdwavecoffee.in", "website": "www.thirdwavecoffee.in", "rating": 4.4, "num_outlets": 65, "decision_maker_name": "Sushant Goel", "decision_maker_role": "Co-founder", "decision_maker_linkedin": "linkedin.com/in/sushant-thirdwave", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "400-700 kg"},
        {"business_name": "Barbeque Nation Bangalore", "segment": "Restaurant", "city": "Bangalore", "state": "Karnataka", "tier": 1, "address": "Residency Road, Bangalore", "phone": "+91-80-4000-1234", "email": "supply@barbequenation.com", "website": "www.barbequenation.com", "rating": 4.1, "num_outlets": 180, "decision_maker_name": "Kayum Dhanani", "decision_maker_role": "CEO", "decision_maker_linkedin": "linkedin.com/in/barbequenation-india", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "5000-8000 kg"},
        {"business_name": "MTR Foods Restaurant", "segment": "Restaurant", "city": "Bangalore", "state": "Karnataka", "tier": 1, "address": "Lalbagh Road, Bangalore", "phone": "+91-80-2222-0022", "email": "supply@mtrfoods.com", "website": "www.mtrfoods.com", "rating": 4.3, "num_outlets": 40, "decision_maker_name": "Sadananda Maiya", "decision_maker_role": "Procurement Head", "decision_maker_linkedin": "linkedin.com/in/mtr-procurement", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "800-1200 kg"},
        {"business_name": "Taj Falaknuma Palace", "segment": "Hotel", "city": "Hyderabad", "state": "Telangana", "tier": 1, "address": "Falaknuma, Hyderabad", "phone": "+91-40-6629-8585", "email": "falaknuma@tajhotels.com", "website": "www.tajhotels.com", "rating": 4.9, "num_outlets": 1, "decision_maker_name": "Zahir Hussain", "decision_maker_role": "Procurement Head", "decision_maker_linkedin": "linkedin.com/in/taj-falaknuma-procurement", "has_dessert_menu": True, "hotel_category": "5-star", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "400-600 kg"},
        {"business_name": "Paradise Restaurant Hyderabad", "segment": "Restaurant", "city": "Hyderabad", "state": "Telangana", "tier": 1, "address": "SD Road, Secunderabad", "phone": "+91-40-2784-7000", "email": "orders@paradiserestaurant.in", "website": "www.paradiserestaurant.in", "rating": 4.4, "num_outlets": 25, "decision_maker_name": "Mohammed Mateen", "decision_maker_role": "Purchase Manager", "decision_maker_linkedin": "", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "600-1000 kg"},
        {"business_name": "ITC Grand Chola Chennai", "segment": "Hotel", "city": "Chennai", "state": "Tamil Nadu", "tier": 1, "address": "63 Mount Road, Guindy, Chennai", "phone": "+91-44-2220-0000", "email": "grandchola.procurement@itchotels.in", "website": "www.itchotels.in", "rating": 4.8, "num_outlets": 3, "decision_maker_name": "Karthik Rajan", "decision_maker_role": "F&B Procurement Director", "decision_maker_linkedin": "linkedin.com/in/karthik-itcchola", "has_dessert_menu": True, "hotel_category": "5-star", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "600-900 kg"},
        {"business_name": "JW Marriott Pune", "segment": "Hotel", "city": "Pune", "state": "Maharashtra", "tier": 1, "address": "Senapati Bapat Road, Pune", "phone": "+91-20-6683-3333", "email": "procurement@marriottpune.com", "website": "www.marriott.com", "rating": 4.6, "num_outlets": 2, "decision_maker_name": "Rohan Desai", "decision_maker_role": "Purchase Manager", "decision_maker_linkedin": "linkedin.com/in/jwmarriott-pune", "has_dessert_menu": True, "hotel_category": "5-star", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "500-700 kg"},
        {"business_name": "Rebel Foods Pune", "segment": "CloudKitchen", "city": "Pune", "state": "Maharashtra", "tier": 1, "address": "Baner, Pune", "phone": "+91-20-6720-1234", "email": "supply@rebelfoods.com", "website": "www.rebelfoods.com", "rating": 3.9, "num_outlets": 120, "decision_maker_name": "Jaydeep Barman", "decision_maker_role": "Co-founder", "decision_maker_linkedin": "linkedin.com/in/jaydeep-rebelfoods", "has_dessert_menu": False, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "1000-2000 kg"},
        {"business_name": "ITC Royal Bengal Kolkata", "segment": "Hotel", "city": "Kolkata", "state": "West Bengal", "tier": 1, "address": "JBS Haldane Avenue, Kolkata", "phone": "+91-33-4455-8000", "email": "royalbengal.procurement@itchotels.in", "website": "www.itchotels.in", "rating": 4.7, "num_outlets": 2, "decision_maker_name": "Debashis Bose", "decision_maker_role": "Procurement Manager", "decision_maker_linkedin": "linkedin.com/in/itc-kolkata-procurement", "has_dessert_menu": True, "hotel_category": "5-star", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "500-800 kg"},
        {"business_name": "Radisson Blu Jaipur", "segment": "Hotel", "city": "Jaipur", "state": "Rajasthan", "tier": 2, "address": "Outer Ring Road, Jaipur", "phone": "+91-141-5100-100", "email": "procurement@radissonblu-jpr.com", "website": "www.radissonblu.com", "rating": 4.3, "num_outlets": 1, "decision_maker_name": "Suresh Sharma", "decision_maker_role": "F&B Manager", "decision_maker_linkedin": "", "has_dessert_menu": True, "hotel_category": "4-star", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "200-400 kg"},
        {"business_name": "LMB Mithai & Restaurant", "segment": "Mithai", "city": "Jaipur", "state": "Rajasthan", "tier": 2, "address": "Johari Bazaar, Jaipur", "phone": "+91-141-2565-844", "email": "orders@lmbhotel.com", "website": "www.lmbhotel.com", "rating": 4.2, "num_outlets": 5, "decision_maker_name": "Rahul Gupta", "decision_maker_role": "Owner", "decision_maker_linkedin": "", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "300-600 kg"},
        {"business_name": "Havmor Ice Cream", "segment": "IceCream", "city": "Ahmedabad", "state": "Gujarat", "tier": 2, "address": "CG Road, Ahmedabad", "phone": "+91-79-2640-1234", "email": "supply@havmor.com", "website": "www.havmor.com", "rating": 4.2, "num_outlets": 200, "decision_maker_name": "Ankit Chona", "decision_maker_role": "MD & CEO", "decision_maker_linkedin": "linkedin.com/in/ankit-chona-havmor", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "4000-7000 kg"},
        {"business_name": "Tunday Kababi Lucknow", "segment": "Restaurant", "city": "Lucknow", "state": "Uttar Pradesh", "tier": 2, "address": "Aminabad, Lucknow", "phone": "+91-522-2610-3456", "email": "procurement@tundaykababi.com", "website": "www.tundaykababi.com", "rating": 4.4, "num_outlets": 8, "decision_maker_name": "Mohammed Usman", "decision_maker_role": "Owner", "decision_maker_linkedin": "", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "150-300 kg"},
        {"business_name": "Royal Catering Lucknow", "segment": "Catering", "city": "Lucknow", "state": "Uttar Pradesh", "tier": 2, "address": "Gomti Nagar, Lucknow", "phone": "+91-522-4012-5678", "email": "info@royalcateringlko.com", "website": "www.royalcateringlko.com", "rating": 4.1, "num_outlets": 1, "decision_maker_name": "Avinash Tripathi", "decision_maker_role": "Owner", "decision_maker_linkedin": "", "has_dessert_menu": True, "hotel_category": "", "is_chain": False, "source": "mock_data", "monthly_volume_estimate": "200-500 kg"},
        {"business_name": "Ratnasagar Sweets Surat", "segment": "Mithai", "city": "Surat", "state": "Gujarat", "tier": 2, "address": "Ring Road, Surat", "phone": "+91-261-2567-890", "email": "procurement@ratnasagar.com", "website": "www.ratnasagarsweets.com", "rating": 4.3, "num_outlets": 12, "decision_maker_name": "Ashish Patel", "decision_maker_role": "Owner", "decision_maker_linkedin": "", "has_dessert_menu": True, "hotel_category": "", "is_chain": True, "source": "mock_data", "monthly_volume_estimate": "400-700 kg"},
        {"business_name": "Pal Dhaba Chandigarh", "segment": "Restaurant", "city": "Chandigarh", "state": "Punjab", "tier": 2, "address": "Sector 28, Chandigarh", "phone": "+91-172-2706-7890", "email": "orders@paldhaba.com", "website": "www.paldhaba.com", "rating": 4.4, "num_outlets": 3, "decision_maker_name": "Kulwant Singh", "decision_maker_role": "Owner", "decision_maker_linkedin": "", "has_dessert_menu": True, "hotel_category": "", "is_chain": False, "source": "mock_data", "monthly_volume_estimate": "150-250 kg"},
    ]

    async with AsyncSessionLocal() as session:
        statuses = ["new", "new", "new", "new", "contacted", "contacted", "qualified", "converted", "lost"]
        created = 0
        for i, lead_data in enumerate(mock_leads):
            score, priority, reasoning = calculate_lead_score(lead_data)
            status = statuses[i % len(statuses)]
            days_ago = random.randint(0, 45)
            ts = datetime.now(timezone.utc) - timedelta(days=days_ago)
            doc = {
                "id": str(uuid.uuid4()),
                **lead_data,
                "ai_score": score,
                "ai_reasoning": reasoning,
                "priority": priority,
                "status": status,
                "created_at": ts,
                "updated_at": ts
            }
            session.add(LeadModel(**doc))
            created += 1

        await session.commit()
        return {"message": f"Seeded {created} HORECA leads", "count": created}


app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)



