"""
app/api/routes/lead.py
─────────────────────────────────────────────────────────────────────────────
All /leads/* endpoints for the HORECA Lead Intelligence API.
─────────────────────────────────────────────────────────────────────────────
"""
import csv
import io
import json
import logging
import uuid
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import desc, func, or_, select

from app.db.session import AsyncSessionLocal
from app.schemas.lead import BulkCreateRequest, DiscoverRequest, LeadCreate, LeadStatusUpdate
from app.utils.scoring import make_lead_obj
from app.utils import model_to_dict
import app.pipelines.stages as ps
from app.db.orm import Lead, City, Segment, OutreachEmail, Contact
from app.agents.prompts.lead_qualify   import lead_qualify_prompt
from app.agents.prompts.lead_email_api import lead_email_api_prompt
from app.core.openai_client import client as openai_client
from app.core.config import OPENAI_MODEL, SERP_API_KEY as _SERP_API_KEY
from app.core.constants import EmailStatus

logger = logging.getLogger(__name__)

lead_router = APIRouter(prefix="/leads")


# ══════════════════════════════════════════════════════════════════════════════
# LIST / FILTER
# ══════════════════════════════════════════════════════════════════════════════

@lead_router.get("")
async def get_leads(
    city: Optional[str] = None,
    segment: Optional[str] = None,
    priority: Optional[str] = None,
    status: Optional[str] = None,
    min_score: Optional[int] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
):
    async with AsyncSessionLocal() as session:
        stmt       = select(Lead)
        count_stmt = select(func.count()).select_from(Lead)

        if city:
            stmt       = stmt.where(Lead.city.ilike(f"%{city}%"))
            count_stmt = count_stmt.where(Lead.city.ilike(f"%{city}%"))
        if segment:
            stmt       = stmt.where(Lead.segment == segment)
            count_stmt = count_stmt.where(Lead.segment == segment)
        if priority:
            stmt       = stmt.where(Lead.priority == priority)
            count_stmt = count_stmt.where(Lead.priority == priority)
        if status:
            stmt       = stmt.where(Lead.status == status)
            count_stmt = count_stmt.where(Lead.status == status)
        if min_score is not None:
            stmt       = stmt.where(Lead.ai_score >= min_score)
            count_stmt = count_stmt.where(Lead.ai_score >= min_score)
        if date_from:
            try:
                df = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                stmt       = stmt.where(Lead.created_at >= df)
                count_stmt = count_stmt.where(Lead.created_at >= df)
            except ValueError:
                pass
        if date_to:
            try:
                # include the entire end day
                dt = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
                stmt       = stmt.where(Lead.created_at <= dt)
                count_stmt = count_stmt.where(Lead.created_at <= dt)
            except ValueError:
                pass
        if search:
            sf = or_(
                Lead.business_name.ilike(f"%{search}%"),
                Lead.city.ilike(f"%{search}%"),
            )
            stmt       = stmt.where(sf)
            count_stmt = count_stmt.where(sf)

        # Only show leads that have at least one contact with an email (email or email_2)
        has_contact = (
            select(Contact.lead_id)
            .where(
                (Contact.lead_id == Lead.id) &
                (
                    (Contact.email.isnot(None) & (Contact.email != "")) |
                    (Contact.email_2.isnot(None) & (Contact.email_2 != ""))
                )
            )
            .exists()
        )
        stmt       = stmt.where(has_contact)
        count_stmt = count_stmt.where(has_contact)

        total = (await session.execute(count_stmt)).scalar() or 0
        lead_rows = (await session.execute(
            stmt.order_by(desc(Lead.created_at)).offset(skip).limit(limit)
        )).scalars().all()
        leads = [model_to_dict(r) for r in lead_rows]

        # Batch-fetch all contacts for the returned leads in one query
        if lead_rows:
            lead_ids = [r.id for r in lead_rows]
            contact_rows = (await session.execute(
                select(Contact).where(Contact.lead_id.in_(lead_ids))
            )).scalars().all()
            contacts_by_lead: dict[str, list] = {}
            for c in contact_rows:
                contacts_by_lead.setdefault(c.lead_id, []).append(c.to_dict())
            for lead in leads:
                lead["contacts"] = contacts_by_lead.get(lead["id"], [])
        else:
            for lead in leads:
                lead["contacts"] = []

    return {"leads": leads, "total": total}


# ══════════════════════════════════════════════════════════════════════════════
# CSV / EXCEL TEMPLATE
# ══════════════════════════════════════════════════════════════════════════════

@lead_router.get("/csv-template")
async def get_csv_template():
    """Generate an Excel template (.xlsx) with dropdown validation for city & segment."""
    async with AsyncSessionLocal() as session:
        city_rows = (await session.execute(
            select(City).where(City.is_active == True).order_by(City.priority.asc(), City.name.asc())
        )).scalars().all()
        seg_rows = (await session.execute(
            select(Segment).where(Segment.is_active == True).order_by(Segment.priority.asc(), Segment.label.asc())
        )).scalars().all()

    city_names = [c.name for c in city_rows] or ["Mumbai", "Delhi", "Bangalore"]
    seg_keys   = [s.key  for s in seg_rows]  or ["Hotel", "Restaurant", "Cafe", "Bakery"]

    wb     = openpyxl.Workbook()
    ref_ws = wb.create_sheet("_ref", 0)
    ref_ws.sheet_state = "hidden"
    for i, name in enumerate(city_names, start=1):
        ref_ws.cell(row=i, column=1, value=name)
    for i, key in enumerate(seg_keys, start=1):
        ref_ws.cell(row=i, column=2, value=key)

    city_ref = f"_ref!$A$1:$A${len(city_names)}"
    seg_ref  = f"_ref!$B$1:$B${len(seg_keys)}"

    ws = wb.create_sheet("Leads", 1)
    wb.active = ws

    COLUMNS = [
        ("business_name", "Business Name *"),   ("segment", "Segment *"),          ("city", "City *"),
        ("state", "State"),                       ("country", "Country"),             ("tier", "Tier (1/2/3)"),
        ("address", "Address"),
        ("phone", "Phone (Business)"),            ("email", "Email (Business)"),
        ("website", "Company Domain / Website"),
        ("description", "Company Description"),
        ("rating", "Rating (0-5)"),               ("num_outlets", "No. of Outlets"),
        # Decision maker → written to contacts table
        ("contact_name",     "Contact Name"),
        ("job_title",        "Job Title"),
        ("linkedin_profile", "LinkedIn Profile"),
        ("work_email",       "Work Email"),
        ("work_email_2",     "Work Email 2"),
        ("mobile",           "Mobile"),
        ("mobile_2",         "Mobile 2"),
        ("has_dessert_menu", "Has Dessert Menu (true/false)"),
        ("hotel_category",   "Hotel Category"),   ("is_chain", "Is Chain (true/false)"),
        ("monthly_volume_estimate", "Monthly Volume Estimate"),
    ]

    SAMPLE = [
        "The Grand Palace Hotel", seg_keys[0], city_names[0],
        city_rows[0].state if city_rows else "", "India", "1",
        "Colaba, Mumbai",
        "+91-9876543210", "procurement@grandhotel.com",
        "grandhotel.com",
        "5-star luxury hotel with multiple restaurants and banquets",
        "4.5", "3",
        "Rajesh Kumar", "Procurement Manager", "linkedin.com/in/rajeshkumar",
        "rajesh@grandhotel.com", "", "+91-9999999999", "",
        "true", "5-star", "false", "500-800 kg",
    ]

    hdr_fill    = PatternFill("solid", fgColor="627F31")
    hdr_font    = Font(bold=True, color="FFFFFF", size=10)
    hdr_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side   = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side)

    for col_idx, (key, label) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        hdr_cell = ws.cell(row=1, column=col_idx, value=label)
        hdr_cell.fill = hdr_fill; hdr_cell.font = hdr_font
        hdr_cell.alignment = hdr_align; hdr_cell.border = thin_border
        sample_cell = ws.cell(row=2, column=col_idx, value=SAMPLE[col_idx - 1])
        sample_cell.alignment = Alignment(vertical="center")
        wide_cols = {"address": 40, "website": 36, "email": 34, "work_email": 34,
                     "work_email_2": 34, "description": 50, "linkedin_profile": 36,
                     "monthly_volume_estimate": 24}
        ws.column_dimensions[col_letter].width = wide_cols.get(key, max(len(label), 18))

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    MAX_ROW = 1001

    seg_col  = next(i for i, (k, _) in enumerate(COLUMNS, 1) if k == "segment")
    city_col = next(i for i, (k, _) in enumerate(COLUMNS, 1) if k == "city")

    dv_seg = DataValidation(type="list", formula1=seg_ref, allow_blank=True,
                            showDropDown=False, showErrorMessage=True,
                            errorTitle="Invalid segment",
                            error="Choose one of the active segments")
    dv_seg.sqref = f"{get_column_letter(seg_col)}2:{get_column_letter(seg_col)}{MAX_ROW}"
    ws.add_data_validation(dv_seg)

    dv_city = DataValidation(type="list", formula1=city_ref, allow_blank=True,
                             showDropDown=False, showErrorMessage=True,
                             errorTitle="Invalid city",
                             error="Choose one of the active target cities")
    dv_city.sqref = f"{get_column_letter(city_col)}2:{get_column_letter(city_col)}{MAX_ROW}"
    ws.add_data_validation(dv_city)

    bool_cols = [k for k, _ in COLUMNS if k in ("has_dessert_menu", "is_chain")]
    dv_bool = DataValidation(type="list", formula1='"true,false"', allow_blank=True,
                             showDropDown=False, showErrorMessage=True,
                             errorTitle="Invalid value", error="Choose true or false")
    for bool_key in bool_cols:
        col_idx    = next(i for i, (k, _) in enumerate(COLUMNS, 1) if k == bool_key)
        col_letter = get_column_letter(col_idx)
        dv_bool.sqref = (
            f"{col_letter}2:{col_letter}{MAX_ROW}" if not dv_bool.sqref
            else f"{dv_bool.sqref} {col_letter}2:{col_letter}{MAX_ROW}"
        )
    ws.add_data_validation(dv_bool)

    tier_col_idx = next(i for i, (k, _) in enumerate(COLUMNS, 1) if k == "tier")
    dv_tier = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True,
                             showDropDown=False, showErrorMessage=True,
                             errorTitle="Invalid tier", error="Choose 1, 2, or 3")
    dv_tier.sqref = f"{get_column_letter(tier_col_idx)}2:{get_column_letter(tier_col_idx)}{MAX_ROW}"
    ws.add_data_validation(dv_tier)

    hcat_col_idx = next((i for i, (k, _) in enumerate(COLUMNS, 1) if k == "hotel_category"), None)
    if hcat_col_idx:
        dv_hcat = DataValidation(type="list", formula1='"3-star,4-star,5-star"', allow_blank=True,
                                 showDropDown=False, showErrorMessage=True,
                                 errorTitle="Invalid category", error="Choose 3-star, 4-star, or 5-star")
        dv_hcat.sqref = f"{get_column_letter(hcat_col_idx)}2:{get_column_letter(hcat_col_idx)}{MAX_ROW}"
        ws.add_data_validation(dv_hcat)

    info_ws = wb.create_sheet("Valid Values", 2)
    info_ws.column_dimensions["A"].width = 24
    info_ws.column_dimensions["B"].width = 24

    def write_section(sheet, start_row, title, values, col):
        hdr = sheet.cell(row=start_row, column=col, value=title)
        hdr.font = Font(bold=True, color="FFFFFF", size=10)
        hdr.fill = PatternFill("solid", fgColor="627F31")
        hdr.alignment = Alignment(horizontal="center")
        for i, v in enumerate(values, start=start_row + 1):
            sheet.cell(row=i, column=col, value=v)

    write_section(info_ws, 1, f"Active Cities ({len(city_names)})", city_names, 1)
    write_section(info_ws, 1, f"Active Segments ({len(seg_keys)})", seg_keys, 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=horeca_leads_template.xlsx"},
    )


# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD CSV
# ══════════════════════════════════════════════════════════════════════════════

@lead_router.post("/upload-csv")
async def upload_csv(file: UploadFile = File(...)):
    content  = await file.read()
    filename = (file.filename or "").lower()
    created  = []
    errors   = []

    # ── Parse into a list of dicts (rows) ────────────────────────────────────
    rows: list[dict] = []

    is_xlsx = (
        filename.endswith(".xlsx")
        or filename.endswith(".xls")
        or file.content_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        )
    )

    if is_xlsx:
        try:
            wb   = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            # use first non-hidden, non-"_ref"/"Valid Values" sheet
            ws = next(
                (wb[name] for name in wb.sheetnames
                 if not name.startswith("_") and name != "Valid Values"),
                wb.active,
            )
            iter_rows = list(ws.iter_rows(values_only=True))
            if not iter_rows:
                return {"created": 0, "errors": ["Empty spreadsheet"]}
            headers = [str(h).strip() if h is not None else "" for h in iter_rows[0]]
            for raw_row in iter_rows[1:]:
                # skip completely empty rows
                if all(v is None or str(v).strip() == "" for v in raw_row):
                    continue
                rows.append({headers[i]: (str(v).strip() if v is not None else "")
                              for i, v in enumerate(raw_row) if i < len(headers)})
            wb.close()
        except Exception as exc:
            return {"created": 0, "errors": [f"Could not parse Excel file: {exc}"]}
    else:
        # CSV — normalise encoding then line endings
        try:
            text = content.decode("utf-8-sig")
        except Exception:
            text = content.decode("latin-1")
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        rows = list(csv.DictReader(io.StringIO(text)))

    # ── Process rows ─────────────────────────────────────────────────────────
    async with AsyncSessionLocal() as session:
        for i, row in enumerate(rows):
            try:
                # ── normalise column aliases from external Excel exports ───────
                # Build a lowercase-keyed lookup so matching is case-insensitive
                row_ci = {k.lower().strip(): v for k, v in row.items()}

                def _col(*keys: str) -> str:
                    for k in keys:
                        v = str(row_ci.get(k.lower(), "")).strip()
                        if v:
                            return v
                    return ""

                raw_domain  = _col("company_domain", "Company Domain", "Company domain", "Company Domain / Website", "website", "Website")
                website_val = (f"https://{raw_domain}" if raw_domain and not raw_domain.startswith("http") else raw_domain)

                first = _col("first_name", "First Name")
                last  = _col("last_name",  "Last Name")
                full  = _col("contact_name", "Contact Name", "decision_maker_name", "Decision Maker Name")
                dm_name     = full or f"{first} {last}".strip()
                dm_role     = _col("job_title", "Job Title", "Job title", "decision_maker_role", "Decision Maker Role")
                dm_linkedin = _col("linkedin_profile", "LinkedIn Profile", "LinkedIn profile", "decision_maker_linkedin")
                dm_email    = _col("work_email", "Work Email", "Work email")
                dm_email2   = _col("work_email_2", "Work Email 2", "Work email 2", "work_email2")
                dm_phone    = _col("mobile", "Mobile")
                dm_phone2   = _col("mobile_2", "Mobile 2", "mobile2")

                # Business phone/email come from dedicated columns; fall back to DM columns only if no DM present
                biz_phone = _col("phone", "Phone (Business)") or (dm_phone if not dm_name else "")
                biz_email = _col("email", "Email (Business)") or (dm_email if not dm_name else "")

                # "segment" column takes priority; then Sub industry (more specific) then Industry (general)
                segment_val = (
                    _col("segment", "Segment *", "Segment",
                         "sub_industry", "Sub industry", "Sub Industry",
                         "industry", "Industry")
                    or "Restaurant"
                )

                # "has_dessert_menu" column (exact lowercase) or legacy label
                has_dessert_raw = _col("has_dessert_menu", "Has Dessert Menu (true/false)")
                has_dessert_val = has_dessert_raw.lower() in ("true", "1", "yes")

                lead_data = {
                    "business_name":           _col("business_name", "Business Name *", "Business Name", "company_name", "Company Name") or "",
                    "segment":                 segment_val,
                    "city":                    _col("city", "City *", "City") or "",
                    "state":                   _col("state", "State") or "",
                    "country":                 _col("country", "Country") or "India",
                    "tier":                    int(_col("tier", "Tier (1/2/3)", "Tier") or "1"),
                    "address":                 _col("address", "Address") or "",
                    "phone":                   biz_phone,
                    "email":                   biz_email,
                    "website":                 website_val,
                    "description":             _col("description", "Company Description", "Company description", "company_description") or "",
                    "rating":                  float(_col("rating", "Rating (0-5)", "Rating") or "0"),
                    "num_outlets":             int(_col("num_outlets", "No. of Outlets") or "1"),
                    "has_dessert_menu":        has_dessert_val,
                    "hotel_category":          _col("hotel_category", "Hotel Category") or "",
                    "is_chain":                _col("is_chain", "Is Chain (true/false)").lower() in ("true", "1", "yes"),
                    "source":                  "csv_upload",
                    "monthly_volume_estimate": _col("monthly_volume_estimate", "Monthly Volume Estimate") or "",
                    # passed through for scoring heuristic only (not written to leads)
                    "decision_maker_name": dm_name,
                }
                if not lead_data["business_name"] or not lead_data["city"]:
                    errors.append(f"Row {i + 2}: Missing business_name or city")
                    continue

                doc = make_lead_obj(lead_data)
                session.add(doc)
                await session.flush()   # populate doc.id before inserting child Contact

                if dm_name:
                    session.add(Contact(
                        lead_id          = doc.id,
                        name             = dm_name,
                        role             = dm_role,
                        email            = dm_email,
                        email_2          = dm_email2,
                        phone            = dm_phone,
                        phone_2          = dm_phone2,
                        linkedin_url     = dm_linkedin,
                        confidence_score = 1.0,
                        source           = "csv_upload",
                        is_primary       = True,
                    ))

                created.append(doc.to_dict())
            except Exception as e:
                errors.append(f"Row {i + 2}: {str(e)}")
        await session.commit()

    return {"created": len(created), "errors": errors}


# ══════════════════════════════════════════════════════════════════════════════
# DISCOVER (pipeline)
# ══════════════════════════════════════════════════════════════════════════════

@lead_router.post("/discover")
async def discover_leads(req: DiscoverRequest):
    city    = req.city.strip()
    segment = req.segment.strip()
    state   = req.state.strip()

    logger.info(f"[DISCOVER] city={city!r}, segment={segment!r}, state={state!r}")

    if _SERP_API_KEY:
        try:
            async with AsyncSessionLocal() as session:
                raw      = await ps.extract_business_data(city, session)
                enriched = await ps.ai_process_business_data(raw, session)
                filtered = await ps.apply_kpi_filtering(enriched, session)
                deduped  = await ps.deduplicate_leads(filtered, session)

            results = [{
                "business_name":           biz.get("business_name", ""),
                "segment":                 biz.get("segment", segment),
                "city":                    biz.get("city", city),
                "state":                   biz.get("state", state),
                "country":                 biz.get("country", "India"),
                "tier":                    biz.get("tier", 1),
                "address":                 biz.get("address", ""),
                "phone":                   biz.get("phone", ""),
                "email":                   biz.get("email", ""),
                "website":                 biz.get("website", ""),
                "description":             biz.get("description", ""),
                "rating":                  biz.get("rating", 0.0),
                "num_outlets":             biz.get("num_outlets", 1),
                "hotel_category":          biz.get("hotel_category", ""),
                "is_chain":                bool(biz.get("is_chain", False)),
                "has_dessert_menu":        bool(biz.get("has_dessert_menu", False)),
                "monthly_volume_estimate": f"{biz.get('monthly_sugar_estimate_kg', '')} kg".strip(),
                "ai_score":                int(biz.get("kpi_score", 0) or 0),
                "ai_reasoning":            biz.get("ai_reasoning", ""),
                "priority":                biz.get("priority", "Low"),
                "source":                  biz.get("source", "serpapi_google_maps"),
                # contacts returned separately for the UI to preview
                "contacts":                biz.get("contacts", []),
            } for biz in deduped]
            logger.info(f"[DISCOVER] Returning {len(results)} leads")
            return results
        except Exception as exc:
            logger.info(f"[DISCOVER] Pipeline failed: {exc}", exc_info=True)
            return []

    logger.info("[DISCOVER] SERP_API_KEY not configured — returning empty results")
    return []


# ══════════════════════════════════════════════════════════════════════════════
# BULK CREATE
# ══════════════════════════════════════════════════════════════════════════════

@lead_router.post("/bulk-create")
async def bulk_create_leads(req: BulkCreateRequest):
    created = []
    async with AsyncSessionLocal() as session:
        for lead_dict in req.leads:
            lead_dict.pop("ai_score", None)
            lead_dict.pop("ai_reasoning", None)
            lead_dict.pop("priority", None)
            doc = make_lead_obj(lead_dict)
            session.add(doc)
            created.append(doc.to_dict())
        await session.commit()
    return {"created": len(created), "leads": created}


# ══════════════════════════════════════════════════════════════════════════════
# CRUD
# ══════════════════════════════════════════════════════════════════════════════

@lead_router.post("")
async def create_lead(lead: LeadCreate):
    doc = make_lead_obj(lead.model_dump())
    async with AsyncSessionLocal() as session:
        session.add(doc)
        await session.commit()
    return doc.to_dict()


@lead_router.get("/{lead_id}")
async def get_lead(lead_id: str):
    async with AsyncSessionLocal() as session:
        lead = (await session.execute(select(Lead).where(Lead.id == lead_id))).scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        lead_dict = model_to_dict(lead)
        contact_rows = (await session.execute(
            select(Contact).where(Contact.lead_id == lead_id)
        )).scalars().all()
        lead_dict["contacts"] = [c.to_dict() for c in contact_rows]
        return lead_dict


@lead_router.put("/{lead_id}/status")
async def update_lead_status(lead_id: str, body: LeadStatusUpdate):
    async with AsyncSessionLocal() as session:
        lead = (await session.execute(select(Lead).where(Lead.id == lead_id))).scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        lead.status     = body.status
        lead.updated_at = datetime.now(timezone.utc)
        await session.commit()
        await session.refresh(lead)
        return model_to_dict(lead)


@lead_router.delete("/{lead_id}")
async def delete_lead(lead_id: str):
    async with AsyncSessionLocal() as session:
        lead = (await session.execute(select(Lead).where(Lead.id == lead_id))).scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        await session.delete(lead)
        await session.commit()
    return {"message": "Lead deleted"}


# ══════════════════════════════════════════════════════════════════════════════
# AI QUALIFY
# ══════════════════════════════════════════════════════════════════════════════

@lead_router.post("/{lead_id}/qualify-ai")
async def qualify_lead_ai(lead_id: str):
    async with AsyncSessionLocal() as session:
        lead_obj = (await session.execute(
            select(Lead).where(Lead.id == lead_id)
        )).scalar_one_or_none()
        if not lead_obj:
            raise HTTPException(status_code=404, detail="Lead not found")
        lead = model_to_dict(lead_obj)

    try:
        prompt = lead_qualify_prompt(
            business_name    = lead['business_name'],
            segment          = lead['segment'],
            city             = lead['city'],
            state            = lead.get('state') or 'India',
            rating           = lead['rating'],
            num_outlets      = lead['num_outlets'],
            hotel_category   = lead.get('hotel_category') or 'N/A',
            has_dessert_menu = lead['has_dessert_menu'],
            is_chain         = lead['is_chain'],
        )

        completion = await openai_client.chat.completions.create(
            model=OPENAI_MODEL,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": "You are a HORECA lead qualification assistant. Always respond with valid JSON."},
                {"role": "user",   "content": prompt},
            ],
        )
        json_str = completion.choices[0].message.content.strip()

        try:
            ai_data = json.loads(json_str)
        except json.JSONDecodeError:
            raise HTTPException(status_code=500, detail="Invalid AI JSON response")

        async with AsyncSessionLocal() as session:
            lead_obj = (await session.execute(
                select(Lead).where(Lead.id == lead_id)
            )).scalar_one_or_none()
            if lead_obj:
                lead_obj.ai_score               = int(ai_data.get("ai_score", lead.get("ai_score", 0)))
                lead_obj.ai_reasoning           = ai_data.get("qualification_summary", "")
                lead_obj.priority               = ai_data.get("priority", "Medium")
                lead_obj.monthly_volume_estimate = ai_data.get("monthly_volume_kg", "")
                lead_obj.updated_at             = datetime.now(timezone.utc)
                await session.commit()
                await session.refresh(lead_obj)
                updated = model_to_dict(lead_obj)

        return {"lead": updated, "ai_analysis": ai_data}

    except Exception as e:
        logger.info(f"AI qualify error: {e}")
        raise HTTPException(status_code=500, detail=f"AI qualification failed: {str(e)}")


# ══════════════════════════════════════════════════════════════════════════════
# GENERATE EMAIL
# ══════════════════════════════════════════════════════════════════════════════

@lead_router.post("/{lead_id}/generate-email")
async def generate_email(lead_id: str):
    async with AsyncSessionLocal() as session:
        lead_obj = (await session.execute(
            select(Lead).where(Lead.id == lead_id)
        )).scalar_one_or_none()
        if not lead_obj:
            raise HTTPException(status_code=404, detail="Lead not found")
        lead = model_to_dict(lead_obj)

        # fetch primary contact — required for personalisation
        primary_contact = (await session.execute(
            select(Contact)
            .where(Contact.lead_id == lead_id, Contact.is_primary == True)
            .limit(1)
        )).scalar_one_or_none()

        if not primary_contact:
            # fall back to any contact before giving up
            primary_contact = (await session.execute(
                select(Contact)
                .where(Contact.lead_id == lead_id)
                .limit(1)
            )).scalar_one_or_none()

        if not primary_contact:
            raise HTTPException(
                status_code=422,
                detail="No contacts found for this lead. Add a contact before generating an email.",
            )

        print(f"Generating email for lead {lead_id} (DM: {primary_contact.name}, role: {primary_contact.role})")

    try:
        dm         = primary_contact.name or "Procurement Manager"
        first_name = dm.split()[0]

        prompt = lead_email_api_prompt(
            business_name           = lead['business_name'],
            segment                 = lead['segment'],
            city                    = lead['city'],
            dm                      = dm,
            first_name              = first_name,
            role                    = primary_contact.role or 'F&B Head',
            rating                  = lead['rating'],
            num_outlets             = lead['num_outlets'],
            has_dessert_menu        = lead['has_dessert_menu'],
            monthly_volume_estimate = lead.get('monthly_volume_estimate') or 'Unknown',
            reasoning               = lead.get('ai_reasoning') or 'N/A',
        )

        completion = await openai_client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {"role": "system", "content": "You are a B2B sales email writer for Dhampur Green, an Indian sugar supplier. Write professional, personalized outreach emails exactly in the format requested."},
                {"role": "user",   "content": prompt},
            ],
        )
        response     = completion.choices[0].message.content
        lines        = response.strip().split("\n")
        subject      = ""
        body_lines   = []
        past_subject = False

        for line in lines:
            if line.startswith("SUBJECT:") and not past_subject:
                subject      = line.replace("SUBJECT:", "").strip()
                past_subject = True
            else:
                body_lines.append(line)

        body = "\n".join(body_lines).strip()
        now  = datetime.now(timezone.utc)

        async with AsyncSessionLocal() as session:
            # Upsert: update the existing draft if one exists, else create new
            existing_draft = (await session.execute(
                select(OutreachEmail)
                .where(
                    OutreachEmail.lead_id == lead_id,
                    OutreachEmail.status  == EmailStatus.DRAFT,
                )
                .order_by(OutreachEmail.generated_at.desc())
                .limit(1)
            )).scalar_one_or_none()

            if existing_draft:
                existing_draft.subject      = subject
                existing_draft.body         = body
                existing_draft.generated_at = now
                await session.commit()
                await session.refresh(existing_draft)
                return model_to_dict(existing_draft)
            else:
                new_email = OutreachEmail(
                    id           = str(uuid.uuid4()),
                    lead_id      = lead_id,
                    lead_name    = lead["business_name"],
                    lead_city    = lead["city"],
                    lead_segment = lead["segment"],
                    subject      = subject,
                    body         = body,
                    status       = EmailStatus.DRAFT,
                    generated_at = now,
                )
                session.add(new_email)
                await session.commit()
                await session.refresh(new_email)
                return model_to_dict(new_email)

    except Exception as e:
        logger.info(f"Email gen error: {e}")
        raise HTTPException(status_code=500, detail=f"Email generation failed: {str(e)}")
